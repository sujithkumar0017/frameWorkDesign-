import openpyxl

def getRowCount(file,sheetName):
    wb = openpyxl.load_workbook(file)
    sheet= wb[sheetName]
    return sheet.max_row
def getColumnCount(file,sheetName):
    wb = openpyxl.load_workbook(file)
    sheet= wb[sheetName]
    return sheet.max_cloumn
def readData(file,sheetName,rownum,columno):
    wb = openpyxl.load_workbook(file)
    sheet= wb[sheetName]
    return sheet.cell(row=rownum,column=columno).value

def writeData(file,sheetName,rownum,columno,data):
    wb = openpyxl.load_workbook(file)
    sheet= wb[sheetName]
    sheet.cell(row=rownum,column=columno).value = data
    wb.save(file)














































